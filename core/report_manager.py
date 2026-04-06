from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import logging

merge_log = logging.getLogger("ReportMerge")


class ReportGenerator:
    def __init__(self, execution_dir: Path):
        self.execution_dir = execution_dir
        self.excel_path = execution_dir / "Long Run Automation report.xlsx"
        self._ensure_workbook()

    def _ensure_workbook(self):
        """Create the Excel workbook if it doesn't exist yet."""
        if not self.excel_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws.cell(row=1, column=1, value="Module-wise test reports — see individual sheets")
            wb.save(self.excel_path)

    def add_module_report(self, module_name, device_id, overall_status, steps,
                          summary_info=None):
        """
        Create / append a dedicated sheet for a test module with one row per
        step, an embedded screenshot thumbnail in each row, and a description
        of what was done.

        Parameters
        ----------
        module_name : str
            Sheet name (e.g. "Parental Lock Setup").  Truncated to 31 chars.
        device_id : str
            Device under test.
        overall_status : str
            "PASSED" or "FAILED".
        steps : list[dict]
            Each dict must have:
                step_number   : int
                step_name     : str   (short label, e.g. "Verify Home Screen")
                description   : str   (what happened in this step)
                status        : str   ("PASSED" / "FAILED" / "SKIPPED")
                screenshot    : str | Path | None  (path to PNG file)
                error_message : str   (empty string if no error)
        summary_info : dict | None
            Optional key-value pairs added at the top of the sheet.
        """
        wb = load_workbook(self.excel_path)

        # Sheet name: max 31 chars, no duplicates
        sheet_title = module_name[:31]
        if sheet_title in wb.sheetnames:
            del wb[sheet_title]
        ws = wb.create_sheet(title=sheet_title)

        # ── Styles ────────────────────────────────────────────────
        header_fill = PatternFill(start_color="00008B", end_color="00008B",
                                  fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                fill_type="solid")
        skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                                fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        wrap_align = Alignment(horizontal="left", vertical="top",
                               wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")

        # ── Summary block (rows 1-N) ─────────────────────────────
        row = 1
        summary_data = {
            "Module": module_name,
            "Device": device_id,
            "Status": overall_status,
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        if summary_info:
            summary_data.update(summary_info)

        for key, val in summary_data.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=1).border = thin_border
            status_cell = ws.cell(row=row, column=2, value=str(val))
            status_cell.border = thin_border
            if key == "Status":
                status_cell.fill = pass_fill if val == "PASSED" else fail_fill
                status_cell.font = Font(bold=True)
            row += 1

        row += 1  # blank row

        # ── Header row ────────────────────────────────────────────
        headers = ["Step #", "Step Name", "Description", "Status",
                    "Screenshot", "Error"]
        col_widths = [8, 28, 55, 12, 32, 40]
        for col_idx, (hdr, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=row, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        row += 1

        # ── Step rows with embedded screenshots ───────────────────
        IMG_HEIGHT_PX = 120          # thumbnail height in the cell
        ROW_HEIGHT_PT = 95           # Excel row height in points

        for step in steps:
            ws.row_dimensions[row].height = ROW_HEIGHT_PT

            # Step #
            c = ws.cell(row=row, column=1, value=step.get("step_number", ""))
            c.alignment = center_align
            c.border = thin_border

            # Step Name
            c = ws.cell(row=row, column=2, value=step.get("step_name", ""))
            c.alignment = wrap_align
            c.border = thin_border

            # Description
            c = ws.cell(row=row, column=3, value=step.get("description", ""))
            c.alignment = wrap_align
            c.border = thin_border

            # Status
            st = step.get("status", "")
            c = ws.cell(row=row, column=4, value=st)
            c.alignment = center_align
            c.border = thin_border
            if st == "PASSED":
                c.fill = pass_fill
            elif st == "FAILED":
                c.fill = fail_fill
            elif st == "SKIPPED":
                c.fill = skip_fill

            # Screenshot — embed as image if file exists
            ss_path = step.get("screenshot")
            if ss_path and Path(ss_path).exists():
                try:
                    img = XlImage(str(ss_path))
                    # Scale to thumbnail keeping aspect ratio
                    ratio = IMG_HEIGHT_PX / img.height if img.height else 1
                    img.width = int(img.width * ratio)
                    img.height = IMG_HEIGHT_PX
                    anchor = f"{get_column_letter(5)}{row}"
                    ws.add_image(img, anchor)
                except Exception as img_exc:
                    ws.cell(row=row, column=5, value=f"(img error: {img_exc})")
            else:
                ws.cell(row=row, column=5, value="—")
            ws.cell(row=row, column=5).border = thin_border

            # Error
            c = ws.cell(row=row, column=6,
                        value=step.get("error_message", ""))
            c.alignment = wrap_align
            c.border = thin_border

            row += 1

        wb.save(self.excel_path)

    def add_locked_channels_sheet(self, channels):
        """
        Add a 'Locked Channels' sheet to the report workbook.

        Parameters
        ----------
        channels : list[dict]
            Each dict: {"channel_number": str, "channel_name": str, "locked": bool}
        """
        wb = load_workbook(self.excel_path)

        sheet_title = "Locked Channels"
        if sheet_title in wb.sheetnames:
            del wb[sheet_title]
        ws = wb.create_sheet(title=sheet_title)

        header_fill = PatternFill(start_color="00008B", end_color="00008B",
                                  fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                fill_type="solid")

        headers = ["S.No", "Channel Number", "Channel Name", "Lock Status"]
        col_widths = [8, 18, 40, 14]
        for col_idx, (hdr, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        for i, ch in enumerate(channels, start=1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=ch.get("channel_number", "?")).border = thin_border
            ws.cell(row=row, column=3, value=ch.get("channel_name", "")).border = thin_border
            status_cell = ws.cell(row=row, column=4,
                                  value="Locked" if ch.get("locked") else "Unlocked")
            status_cell.alignment = center_align
            status_cell.border = thin_border
            if ch.get("locked"):
                status_cell.fill = pass_fill

        wb.save(self.excel_path)

    def add_favourite_channels_sheet(self, channels):
        """
        Add a 'Favourite Channels' sheet to the report workbook.

        Parameters
        ----------
        channels : list[dict]
            Each dict: {"channel_number": str, "channel_name": str, "selected": bool}
        """
        wb = load_workbook(self.excel_path)

        sheet_title = "Favourite Channels"
        if sheet_title in wb.sheetnames:
            del wb[sheet_title]
        ws = wb.create_sheet(title=sheet_title)

        header_fill = PatternFill(start_color="00008B", end_color="00008B",
                                  fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                fill_type="solid")

        headers = ["S.No", "Channel Number", "Channel Name", "Status"]
        col_widths = [8, 18, 40, 14]
        for col_idx, (hdr, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        for i, ch in enumerate(channels, start=1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=ch.get("channel_number", "?")).border = thin_border
            ws.cell(row=row, column=3, value=ch.get("channel_name", "")).border = thin_border
            status_cell = ws.cell(row=row, column=4,
                                  value="Selected ★" if ch.get("selected") else "Not selected")
            status_cell.alignment = center_align
            status_cell.border = thin_border
            if ch.get("selected"):
                status_cell.fill = pass_fill

        wb.save(self.excel_path)

    @staticmethod
    def merge_device_reports(execution_root_dir):
        """
        Merge all device-specific Excel reports into one combined report
        at the execution root level.

        Parameters
        ----------
        execution_root_dir : str | Path
            Path to the execution root, e.g. TestResults/Execution_20260406_124651
        """
        execution_root = Path(execution_root_dir)
        combined_path = execution_root / "Long Run Automation report.xlsx"

        # Find all device-specific Excel files
        device_excels = sorted(
            execution_root.glob("device_*/Long Run Automation report.xlsx")
        )

        if not device_excels:
            merge_log.warning(f"No device Excel reports found in {execution_root}")
            return None

        # If only one device, just copy it
        if len(device_excels) == 1:
            import shutil
            shutil.copy2(device_excels[0], combined_path)
            merge_log.info(f"Single device report copied to {combined_path}")
            return combined_path

        # Create combined workbook
        combined_wb = Workbook()
        summary_ws = combined_wb.active
        summary_ws.title = "Summary"

        # Build summary header
        header_fill = PatternFill(start_color="00008B", end_color="00008B",
                                  fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        center_align = Alignment(horizontal="center", vertical="center")

        summary_headers = ["Device", "Sheets Found", "Status"]
        for col_idx, hdr in enumerate(summary_headers, 1):
            cell = summary_ws.cell(row=1, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 40
        summary_ws.column_dimensions['C'].width = 15

        summary_row = 2

        for device_excel in device_excels:
            device_folder = device_excel.parent.name  # e.g. device_172_18_1_98_5555
            short_device = device_folder.replace("device_", "")

            try:
                wb = load_workbook(device_excel)
                sheet_names = [s for s in wb.sheetnames if s != "Summary"]

                # Summary row for this device
                summary_ws.cell(row=summary_row, column=1,
                                value=short_device).border = thin_border
                summary_ws.cell(row=summary_row, column=2,
                                value=", ".join(sheet_names)).border = thin_border
                summary_ws.cell(row=summary_row, column=3,
                                value="Merged").border = thin_border
                summary_row += 1

                for sheet_name in sheet_names:
                    src_ws = wb[sheet_name]

                    # Build unique sheet name: "DeviceShort_Module" (max 31 chars)
                    # Use last octet of IP for brevity
                    device_tag = short_device.split("_")[-2] if "_" in short_device else short_device[:8]
                    new_name = f"{sheet_name[:22]}_{device_tag}"[:31]

                    # Handle duplicate sheet names
                    base_new_name = new_name
                    counter = 2
                    while new_name in combined_wb.sheetnames:
                        new_name = f"{base_new_name[:28]}_{counter}"[:31]
                        counter += 1

                    dst_ws = combined_wb.create_sheet(title=new_name)

                    # Copy cell values, formatting, and dimensions
                    for row in src_ws.iter_rows():
                        for cell in row:
                            new_cell = dst_ws.cell(
                                row=cell.row, column=cell.column,
                                value=cell.value
                            )
                            if cell.has_style:
                                new_cell.font = cell.font.copy()
                                new_cell.fill = cell.fill.copy()
                                new_cell.alignment = cell.alignment.copy()
                                new_cell.border = cell.border.copy()
                                new_cell.number_format = cell.number_format

                    # Copy column widths
                    for col_letter, dim in src_ws.column_dimensions.items():
                        dst_ws.column_dimensions[col_letter].width = dim.width

                    # Copy row heights
                    for row_num, dim in src_ws.row_dimensions.items():
                        dst_ws.row_dimensions[row_num].height = dim.height

                    # Copy images/screenshots
                    for img in src_ws._images:
                        try:
                            new_img = XlImage(img.ref)
                            new_img.width = img.width
                            new_img.height = img.height
                            new_img.anchor = img.anchor
                            dst_ws.add_image(new_img)
                        except Exception:
                            pass  # skip images that can't be copied

                wb.close()
                merge_log.info(f"Merged {len(sheet_names)} sheet(s) from {device_folder}")

            except Exception as e:
                merge_log.warning(f"Failed to merge {device_excel}: {e}")
                summary_ws.cell(row=summary_row - 1, column=3,
                                value=f"ERROR: {e}").border = thin_border

        combined_wb.save(combined_path)
        merge_log.info(f"Combined report saved: {combined_path}")
        return combined_path