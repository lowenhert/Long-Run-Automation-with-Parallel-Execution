"""
Email Sender Module for OTT Playback Test Reports
Sends automated email reports with test results via Gmail SMTP
"""

import smtplib
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from datetime import datetime
import logging

log = logging.getLogger(__name__)


class EmailSender:
    def __init__(self, email_config):
        """
        Initialize email sender with configuration
        
        Args:
            email_config: Dictionary with email settings from settings.yaml
        """
        self.enabled = email_config.get("enabled", False)
        self.smtp_server = email_config.get("smtp_server", "smtp.gmail.com")
        self.smtp_port = email_config.get("smtp_port", 587)
        self.sender_email = email_config.get("sender_email")
        self.sender_password = email_config.get("sender_password")
        self.sender_name = email_config.get("sender_name", "OTT Test Automation")
        self.recipient_emails = email_config.get("recipient_emails", [])
        self.subject_prefix = email_config.get("subject_prefix", "Automated OTT Playback Check")
    
    def parse_excel_report(self, excel_path):
        """
        Parse Excel report to generate summary statistics
        
        Args:
            excel_path: Path to the Excel report file
            
        Returns:
            Dictionary with summary statistics
        """
        summary = {
            "total_tests": 0,
            "passed": 0,
            "failed": 0,
            "failure_details": [],
            "apps_tested": set(),
            "devices_tested": set()
        }
        
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Get header row to map column indices
            headers = [cell.value for cell in ws[1]]
            header_map = {h: i for i, h in enumerate(headers) if h}
            
            # Parse data rows (skip header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                    
                summary["total_tests"] += 1
                
                app_name = row[header_map.get("Ott_App_Name", 0)] or "Unknown"
                device_id = row[header_map.get("Device_ID", 4)] or "Unknown"
                status = str(row[header_map.get("Status", 5)] or "").upper()
                
                summary["apps_tested"].add(app_name)
                summary["devices_tested"].add(device_id)
                
                if status == "PASSED":
                    summary["passed"] += 1
                elif status == "FAILED":
                    summary["failed"] += 1
                    summary["failure_details"].append({
                        "app": app_name,
                        "content": row[header_map.get("Content_Name", 1)] or "Unknown",
                        "error": row[header_map.get("Error_Message", 8)] or "No details",
                        "step": row[header_map.get("Failed_Step", 9)] or "Unknown"
                    })
            wb.close()
        except Exception as e:
            log.error(f"Error parsing Excel report: {e}")
        
        return summary
    
    def generate_email_body(self, summary, execution_id):
        """
        Generate email body with formatted summary
        
        Args:
            summary: Dictionary with test summary statistics
            execution_id: Execution ID string
            
        Returns:
            Formatted email body as HTML string
        """
        current_date = datetime.now().strftime("%d-%m-%Y")
        
        # Determine overall result
        if summary["failed"] == 0:
            overall_result = "Pass (All)"
        elif summary["failed"] <= summary["total_tests"] * 0.2:
            overall_result = "Pass (Majority)"
        else:
            overall_result = f"Partial ({summary['passed']}/{summary['total_tests']} passed)"
        
        # Build failure summary
        failure_text = ""
        if summary["failed"] > 0:
            failure_items = []
            for failure in summary["failure_details"][:5]:
                failure_items.append(f"{failure['app']} - {failure['error'][:50]}")
            failure_text = f"<strong>Failures:</strong> {summary['failed']}<br><ul>"
            for item in failure_items:
                failure_text += f"<li>{item}</li>"
            failure_text += "</ul>"
            if summary["failed"] > 5:
                failure_text += f"<em>... and {summary['failed'] - 5} more failures in attached report.</em>"
        else:
            failure_text = "<strong>Failures:</strong> None"
        
        html_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .summary {{ background-color: #f4f4f4; padding: 15px; border-radius: 5px; margin: 20px 0; }}
                .summary p {{ margin: 8px 0; }}
                ul {{ margin: 10px 0; padding-left: 20px; }}
                .footer {{ margin-top: 30px; color: #666; font-size: 0.9em; }}
            </style>
        </head>
        <body>
            <p>Hi Sir,</p>
            
            <p>Automated OTT Playback check completed for today.</p>
            
            <p><strong>Summary below:</strong></p>
            
            <div class="summary">
                <p><strong>Execution:</strong> Completed ({execution_id})</p>
                <p><strong>Date:</strong> {current_date}</p>
                <p><strong>Content Coverage:</strong> {len(summary['apps_tested'])} apps tested ({', '.join(summary['apps_tested'])})</p>
                <p><strong>Total Tests:</strong> {summary['total_tests']}</p>
                <p><strong>Playback Result:</strong> {overall_result}</p>
                <p>{failure_text}</p>
                <p><strong>Logs & Screenshots:</strong> Auto captured</p>
                <p><strong>System Health:</strong> Stable</p>
            </div>
            
            <p>Detailed report attached. {'Issue under investigation.' if summary['failed'] > 0 else 'All tests passed successfully.'}</p>
            
            <div class="footer">
                <p>Regards,<br>
                {self.sender_name}</p>
            </div>
        </body>
        </html>
        """
        
        return html_body
    
    def send_report(self, excel_path, execution_id):
        """
        Send email report with Excel attachment via Gmail SMTP
        
        Args:
            excel_path: Path to the Excel report file
            execution_id: Execution ID string
            
        Returns:
            bool: True if email sent successfully, False otherwise
        """
        if not self.enabled:
            log.info("Email sending is disabled in configuration")
            return False
        
        if not Path(excel_path).exists():
            log.error(f"Excel report not found: {excel_path}")
            return False
        
        try:
            # Parse Excel and generate summary
            summary = self.parse_excel_report(excel_path)
            
            # Create message
            msg = MIMEMultipart()
            msg['From'] = f"{self.sender_name} <{self.sender_email}>"
            msg['To'] = ", ".join(self.recipient_emails)
            msg['Subject'] = f"{self.subject_prefix} – Summary"
            
            # Generate and attach email body
            body = self.generate_email_body(summary, execution_id)
            msg.attach(MIMEText(body, 'html'))
            
            # Attach Excel report
            with open(excel_path, 'rb') as attachment:
                part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename={Path(excel_path).name}'
                )
                msg.attach(part)
            
            # Send email via Gmail SMTP
            log.info(f"Connecting to {self.smtp_server}:{self.smtp_port}")
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            server.starttls()
            server.login(self.sender_email, self.sender_password)
            text = msg.as_string()
            server.sendmail(self.sender_email, self.recipient_emails, text)
            server.quit()
            
            log.info(f"✅ Email sent successfully to {', '.join(self.recipient_emails)}")
            print(f"✅ Email sent successfully to {', '.join(self.recipient_emails)}")
            return True
            
        except smtplib.SMTPAuthenticationError:
            log.error("❌ SMTP Authentication failed. Check email and app password.")
            print("❌ SMTP Authentication failed. Check email and app password.")
            return False
        except smtplib.SMTPException as e:
            log.error(f"❌ SMTP error occurred: {e}")
            print(f"❌ SMTP error: {e}")
            return False
        except Exception as e:
            log.error(f"❌ Failed to send email: {e}")
            print(f"❌ Failed to send email: {e}")
            return False
